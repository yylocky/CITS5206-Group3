from flask import render_template, request, redirect, url_for, flash, jsonify, current_app, session
from flask_login import current_user, login_user, logout_user, login_required
from app import app, db, forms
from app.models import Login, User, WorkloadAllocation, Role, Work, Department
from app.forms import LoginForm, SignupForm
from werkzeug.urls import url_parse
from openpyxl import load_workbook
import re
import random
from flask import g
from flask import session

import uuid

# decorator for login page


@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    form = LoginForm()
    if form.validate_on_submit():
        user = Login.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Invalid Username or Password')
            return redirect(url_for('login'))
        login_user(user, remember=form.remember_me.data)
        next_page = request.args.get('next')
        if not next_page or url_parse(next_page).netloc != '':
            next_page = url_for('login')
        set_session(user)
        return redirect(next_page)
    return render_template('login.html', title='Sign In', form=form)

# decorator for logout page


@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))

# decorator for filtered view_workload page from notification


@app.route('/view_workload')
@login_required
def view_workload():
    # Check if the alloc_id is present in the request's query parameters
    alloc_id = request.args.get('alloc_id')

    if alloc_id:
        # Fetch workload entries with the specified alloc_id
        workloads = WorkloadAllocation.query.filter_by(
            alloc_id=alloc_id).all()
    else:
        # If no alloc_id is specified (i.e. accessed from nav bar), fetch all workload entries
        workloads = WorkloadAllocation.query.all()

    return render_template('view_workload.html', title='View Workload', workloads=workloads)

# decorator for view all workload page


@app.route('/view_all_workload')
@login_required
def view_all_workload():
    role_id = current_user.user.role_id
    # if current user is admin(4) or hod(1), show all workload
    if role_id == 4 or role_id == 1:
        workloads = WorkloadAllocation.query.all()

    # if current user is HoD(2), show only their department workload
    elif role_id == 2:
        workloads = WorkloadAllocation.query.join(Work, Work.work_id == WorkloadAllocation.work_id).filter(
            Work.dept_id == current_user.user.dept_id).all()

    # if current user is staff(3), show only their workload
    elif role_id == 3:
        workloads = WorkloadAllocation.query.filter_by(
            username=current_user.username).all()

    role_name = get_session()
    return render_template('view_all_workload.html', title='View All Workload', workloads=workloads, role_name=role_name)


@app.route('/assign_workload')
@login_required
def assign():
    # Check if the currently logged-in user has a role_id of 3
    if current_user.user.role_id == 3:
        flash('You do not have permission to access this page.')
        return redirect(url_for('dashboard'))

    # Fetch all usernames from the User table
    users = User.query.with_entities(User.username).all()
    usernames = [user.username for user in users]
    return render_template('assign_workload.html', title='Assign Workload', usernames=usernames)


@app.route('/edit_allocation_detail', methods=['GET', 'POST'])
@login_required
def edit_allocation_detail():
    role_name = get_session()
    if request.method == 'POST':
        code = 0
        msg = 'error'

        alloc_id = request.form.get('alloc_id')
        comment = request.form.get('comment')
        if comment is None or comment == '':
            msg = 'Invalid Comment'
            rest = {'code': code, 'msg': msg}
            return rest

        info = WorkloadAllocation.query.filter_by(alloc_id=alloc_id).first()
        info.comment = comment

        #update comment status to unread
        info.comment_status = "Unread"
        
        db.session.commit()
        if info:
            msg = 'Comment Success'
        rest = {'code': code, 'msg': msg}
        return rest

    workloads = None
    role_id = session.get('role_id', '')
    username = session.get('username', '')
    user = User.query.filter_by(username=username, role_id=role_id).first()
    # if current user is admin(4) or hod(1), show all workload
    if role_id == 4 or role_id == 1:
        workloads = WorkloadAllocation.query.all()

    # if current user is HoD(2), show only their department workload
    elif role_id == 2:
        workloads = WorkloadAllocation.query.join(Work, Work.work_id == WorkloadAllocation.work_id).filter(
            Work.dept_id == user.dept_id).all()

    # if current user is staff(3), show only their workload
    elif role_id == 3:
        workloads = WorkloadAllocation.query.filter_by(username=username).all()

    data = None
    if request.args.get('alloc_id') is not None:
        data = WorkloadAllocation.query.filter_by(alloc_id=request.args.get('alloc_id')).first()
    return render_template('edit_allocation_detail.html', title='Edit Allocation Detail', role_name=role_name,
                           workloads=workloads, user=user, data=data)


@app.route('/get_work_type', methods=['GET', 'POST'])
@login_required
def get_work_type():
    work_type = []
    work_category = request.form.get('work_category')
    if work_category == "Service":
        work_type = ["ADMIN", "GA", "SDS"]

    if work_category == "Teaching":
        work_type = ["CWS", "TEACH", "UDEV"]

    if work_category == "Research":
        work_type = ["HDR", "ORES", "RES-MGMT", "RESERV"]

    if work_category == "Leave":
        work_type = ["LSL", "PL", "SBL"]

    return work_type


@app.route('/get_work', methods=['GET', 'POST'])
@login_required
def get_work():
    work_id = request.form.get('work_id')
    work = Work.query.filter_by(work_id=work_id).first()
    rest = {
        'work_id': work_id,
        'dept_id': work.dept_id,
        'work_explanation': work.work_explanation,
        'work_type': work.work_type,
        'unit_code': work.unit_code
    }
    return rest


@app.route('/get_works', methods=['GET', 'POST'])
@login_required
def get_works():
    works = Work.query.all()
    rest = []
    for item in works:
        temp = {
            'work_id': item.work_id,
            'dept_id': item.dept_id,
            'work_explanation': item.work_explanation,
            'work_type': item.work_type,
            'unit_code': item.unit_code
        }
        rest.append(temp)

    return rest

@app.route("/get_department", methods=["POST"], endpoint="get_department_view")
@login_required
def get_department():


#@app.route('/get_department', methods=['GET', 'POST'])
#@login_required
#def get_department():
    dept_id = request.form.get('dept_id')
    if dept_id is not None:
        department = Department.query.filter_by(dept_id=dept_id).first()
        rest = {
            'dept_id': department.dept_id,
            'dept_name': department.dept_name,
        }
        return rest

    departments = Department.query.all()
    rest = []
    for item in departments:
        temp = {
            'dept_id': item.dept_id,
            'dept_name': item.dept_name
        }
        rest.append(temp)
    return rest


@app.route('/get_user', methods=['GET', 'POST'])
@login_required
def get_user():
    users = User.query.all()
    rest = []
    for item in users:
        temp = {
            'username': item.username,
            'role_id': item.role_id,
            'leave_hours': item.leave_hours,
            'contract_hour': item.contract_hour,
            'available_hours': item.available_hours,
            'dept_id': item.dept_id,
        }
        rest.append(temp)
    return rest


@app.route('/set_workload_allocation', methods=['GET', 'POST'])
@login_required
def set_workload_allocation():
    code = 0
    msg = "Edit Error"
    alloc_id = request.form.get('alloc_id')
    hours_allocated = request.form.get('hours_allocated')
    workload_point = request.form.get('workload_point')
    if hours_allocated is None or hours_allocated == "":
        rest = {'code': code, 'msg': "Adjusted Hours Is Required"}
        return rest

    if workload_point is None or workload_point == "":
        rest = {'code': code, 'msg': "Adjusted Workload Point Is Required"}
        return rest

    workload = WorkloadAllocation.query.filter_by(alloc_id=alloc_id).first()
    workload.hours_allocated = hours_allocated
    workload.workload_point = workload_point

    db.session.commit()
    if workload:
        code = 1
        msg = "Edit successfully"

    rest = {'code': code, 'msg': msg}
    return rest


def set_session(user):
    user_info = User.query.filter_by(username=user.username).first()
    if user_info is None:
        flash('Invalid Username or Password')
        return redirect(url_for('login'))
    role = Role.query.filter_by(role_id=user_info.role_id).first()
    session['role_id'] = role.role_id
    session['role_name'] = role.role_name
    session['username'] = user.username


def get_session():
    role_name = session.get('role_name', '')
    return role_name


def set_session(user):
    user_info = User.query.filter_by(username=user.username).first()
    if user_info is None:
        flash('Invalid Username or Password')
        return redirect(url_for('login'))
    role = Role.query.filter_by(role_id=user_info.role_id).first()
    session['role_id'] = role.role_id
    session['role_name'] = role.role_name


def get_session():
    role_name = session.get('role_name', '')
    return role_name


@app.route('/dashboard')
@login_required
def dashboard():
    role_id = current_user.user.role_id
    # if current user is admin(4) or hod(1), show all workload
    if role_id == 4 or role_id == 1:
        workloads = WorkloadAllocation.query.filter_by(username=current_user.username).all()

    # if current user is HoD(2), show only their department workload
    elif role_id == 2:
        workloads = WorkloadAllocation.query.join(Work, Work.work_id == WorkloadAllocation.work_id).filter(
            Work.dept_id == current_user.user.dept_id).filter(WorkloadAllocation.username == current_user.username).all()

    # if current user is staff(3), show only their workload
    elif role_id == 3:
        workloads = WorkloadAllocation.query.filter_by(
            username=current_user.username).all()


    deprest = Department.query.all()
    departments = []
    departments_value = []
    for item in deprest:
        value = 0
        departments.append(item.dept_name)

        works = Work.query.filter_by(dept_id=item.dept_id).all()
        for subitem in works:
            workallcation = WorkloadAllocation.query.filter_by(work_id=subitem.work_id).filter_by(username=current_user.username).all()
            for witem in workallcation:
                value = value + witem.hours_allocated
        departments_value.append(value)

    all_work = Work.query.all()
    work_type = []
    work_type_value = []
    work_category = []
    work_category_value = []
    for item in all_work:
        if item.work_type not in work_type:
            work_type.append(item.work_type)

        if item.work_category not in work_category:
            work_category.append(item.work_category)

    for item in work_type:
        value = 0
        work = Work.query.filter_by(work_type=item).all()
        for subitem in work:
            workallcation = WorkloadAllocation.query.filter_by(work_id=subitem.work_id).filter_by(username=current_user.username).all()
            for witem in workallcation:
                value = value + witem.hours_allocated
        work_type_value.append(value)

    work_type_data = []
    i = 0
    for item in work_type:
        temp = {
            'name': item,
            'value': work_type_value[i]
        }
        work_type_data.append(temp)
        i = i + 1

    for item in work_category:
        work_type = []
        if item == "Service":
            work_type = ["ADMIN", "GA", "SDS"]

        if item == "Teaching":
            work_type = ["CWS", "TEACH", "UDEV"]

        if item == "Research":
            work_type = ["HDR", "ORES", "RES-MGMT", "RESERV"]

        if item == "Leave":
            work_type = ["LSL", "PL", "SBL"]
        value = 0
        for subitem in work_type:
            work = Work.query.filter_by(work_type=subitem).all()
            for subitem in work:
                workallcation = WorkloadAllocation.query.filter_by(work_id=subitem.work_id).filter_by(
                    username=current_user.username).all()
                for witem in workallcation:
                    value = value + witem.hours_allocated

        work_category_value.append(value)

    work_category_data = []
    j = 0
    for item in work_category:
        temp = {
            'name': item,
            'value': work_category_value[j]
        }
        work_category_data.append(temp)
        j = j + 1



    return render_template('dashboard.html', title='Dashboard',
                           username=current_user.username,
                           workloads=workloads,
                           departments=departments,
                           departments_value=departments_value,
                           work_type_data=work_type_data,
                           work_category_data=work_category_data
                           )


# Single-line Assignment
@app.route("/getpoint", methods=["POST"])
def get_point():
    staff_id = request.form.get("staffId")

    user = User.query.filter_by(username=staff_id).first()

    if user:
        contract_hour = user.contract_hour

        return jsonify({"contract_hour": contract_hour})
    else:
        return jsonify({"message": "User not found"})


@app.route('/assign', methods=['POST'])
def assign_task():
    try:
        k_category = request.form.get("category")
        k_taskType = request.form.get("taskType")
        k_department = request.form.get("department")
        k_work_id = request.form.get("staffId")
        k_username = k_work_id
        k_comment = request.form.get("explanation")
        k_comment_status = ""
        k_unit_code = request.form.get("unitCode")
        k_hours_allocated = request.form.get("assignedHours")
        k_workload_point = request.form.get("workPoint")

        print("k_category: ")
        print(k_category)
        print("k_taskType: ")
        print(k_taskType)
        print("k_department: ")
        print(k_department)
        print("k_work_id: ")
        print(k_work_id)
        print("k_username: ")
        print(k_username)
        print("k_comment: ")
        print(k_comment)
        print("k_comment_status: ")
        print(k_comment_status)
        print("k_unit_code: ")
        print(k_unit_code)
        print("k_hours_allocated: ")
        print(k_hours_allocated)
        print("k_workload_point: ")
        print(k_workload_point)

        k_depart = Department.query.filter_by(dept_name=k_department).first()
        print(k_depart)
        k_dept_id = k_depart.dept_id
        print(k_dept_id)

        role_name = "HoS"
        k_role = Role.query.filter_by(role_name=role_name).first()
        k_role_id = k_role.role_id
        print(k_role_id)

        # #work
        k_work_explanation = "Some explanation for " + k_taskType
        k_work = Work(
            work_explanation=k_comment,
            work_type=k_taskType,
            dept_id=int(k_dept_id),
            unit_code=k_unit_code
        )
        #
        db.session.add(k_work)
        db.session.commit()

        # #workload
        k_workload_allocation = WorkloadAllocation(
            work_id=k_work.work_id,
            hours_allocated=float(k_hours_allocated),
            username=int(k_username),
            comment_status=k_comment_status,
            workload_point=float(k_workload_point)
        )

        db.session.add(k_workload_allocation)
        db.session.commit()
        #

        # user
        k_contract_hour = float(k_workload_point) / float(k_hours_allocated)
        k_user = User.query.filter_by(username=k_username).first()
        k_user.leave_hours += float(k_hours_allocated)
        db.session.add(k_user)
        db.session.commit()

        return "assigned and data stored successfully."
    except Exception as e:
        return f"Error: {str(e)}"


def generate_int_id():
    # Generate a UUID4 and convert it to an integer
    uuid_str = str(uuid.uuid4()).replace('-', '')
    int_id = int(uuid_str, 16)
    int_id = int(str(int_id)[0:12])
    return int_id


@app.route("/get_department", methods=["POST"])
def get_department():
    staff_id = request.form.get("staffId")
    user_info = User.query.filter_by(username=staff_id).first()
    dept_info = Department.query.filter_by(dept_id=user_info.dept_id).first()
    # Assuming you have retrieved the department as "department_name"
    # Replace with your actual department retrieval logic
    department_name = dept_info.dept_name
    return jsonify({"department": department_name})


@app.route('/check_duplicate_task', methods=['POST'])
def check_duplicate_task():
    try:
        task_type = request.form.get("taskType")
        department = request.form.get("department")
        staff_id = request.form.get("staffId")
        assigned_hours = request.form.get("assignedHours")
        k_workload_point = request.form.get("workPoint")

        is_duplicate = False

        workloads = WorkloadAllocation.query.filter_by(
            username=staff_id, hours_allocated=assigned_hours, workload_point=k_workload_point).all()
        for item in workloads:
            k_depart = Department.query.filter_by(dept_name=department).first()
            works = Work.query.filter_by(
                work_id=item.work_id, work_type=task_type, dept_id=k_depart.dept_id).all()
            if len(works) != 0:
                is_duplicate = True
                break
        print(is_duplicate)
        return jsonify({"is_duplicate": is_duplicate})
    except Exception as e:
        return jsonify({"error": str(e)})


# Upload function, including validate file type - MW
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv', 'tsv'}  # MW

@app.route("/upload", methods=["POST"])
@login_required
def upload_file():
    role_id = current_user.user.role_id
    role = Role.query.filter_by(role_id=role_id).first()
    if role.role_name == "Staff":
        flash('Staff cannot use')
        return redirect(url_for('assign'))

    file = request.files["file"]
    file_extension = file.filename.rsplit(
        '.', 1)[1].lower() if '.' in file.filename else ''  # MW

    if file_extension not in ALLOWED_EXTENSIONS:  # MW
        # return "Invalid file type. Please upload a valid Excel file." #MW
        flash('Invalid file type, please upload again')  # MW
        return redirect(url_for('assign'))  # mw

    if file.filename != "":
        try:

            sheetname = 'Work Assignment'
            workbook = load_workbook(file, read_only=False, data_only=True)
            sheet = workbook[sheetname]

            sheet_title = sheet.title

            if sheet_title != sheetname:
                # return "This may not be the right spreadsheet as it does not pass the content validation."
                flash('Invalid spreedsheet, please upload again')
                return redirect(url_for('assign'))

            data = []
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                data.append(row_data)

            workbook.close()
            keys = [
                "Staff Number",
                "Type",
                "Department",
                "Unit Code",
                "Explanation",
                "Assigned Hours",
            ]

            rest = []

            for item in data[1:]:
                temp = dict(zip(keys, item))
                rest.append(temp)

            for item in rest:
                if item['Staff Number'] is None:
                    continue

                if item['Department'] is None:
                    continue

                user = User.query.filter_by(
                    username=item['Staff Number']).first()
                if user is None:
                    flash("User {} does not exist.  All workloads before the invalid entry have been successfully assigned.".format(
                        item['Staff Number']))
                    return redirect(url_for('assign'))

                if user.contract_hour is None or user.contract_hour == 0:
                    flash("User {} contract_hour is empty.".format(
                        item['Staff Number']))
                    return redirect(url_for('assign'))

                dept = Department.query.filter_by(
                    dept_name=item['Department']).first()
                if dept is None:
                    flash("Department {} does not exist.".format(
                        item['Department']))
                    return redirect(url_for('assign'))

                # data insert work
                if item['Unit Code'] is not None:
                    work_explanation = item['Explanation']
                    work_type = item['Type']
                    dept_id = dept.dept_id
                    unit_code = item['Unit Code']

                    work = Work.query.filter_by(
                        work_explanation=work_explanation,
                        work_type=work_type,
                        dept_id=dept_id,
                        unit_code=unit_code,
                    ).first()

                    if work is None:
                        work = Work(
                            work_explanation=work_explanation,
                            work_type=work_type,
                            dept_id=dept_id,
                            unit_code=unit_code,
                        )
                        db.session.add(work)
                        db.session.commit()

                    # data insert workload_allocation
                    work_id = work.work_id
                    hours_allocated = 0
                    if item['Assigned Hours'] is not None:
                        hours_allocated = item['Assigned Hours']

                    typeArr = [
                        "PL",
                        "SBL",
                        "LSL",
                    ]

                    if item['Type'] in typeArr:
                        user.leave_hours += hours_allocated
                        db.session.commit()

                    username = item['Staff Number']
                    comment = ''
                    comment_status = ''
                    workload_point = 0
                    if hours_allocated != 0:
                        workload_point = round(
                            float(hours_allocated) / user.contract_hour, 2)

                    workload_allocation = WorkloadAllocation(
                        work_id=work_id,
                        hours_allocated=hours_allocated,
                        username=username,
                        comment=comment,
                        comment_status=comment_status,
                        workload_point=workload_point,
                    )
                    db.session.add(workload_allocation)
                    db.session.commit()

            flash("File uploaded and data stored as TaskData objects successfully.")
            return redirect(url_for('assign'))
        except Exception as e:
            return f"Error: {str(e)}"

    flash("No file selected for upload.")
    return redirect(url_for('assign'))


@app.route('/comment_history')
@login_required
def comment_history():
    # Check if the current user is an HoD
    if g.is_hod:
        # Fetch the department ID of the current HoD user
        hod_dept_id = User.query.filter_by(
            username=current_user.username).first().dept_id

        # Fetch the comments made by staff members with the same dept_id as that of the HoD user and save them in the comments variable
        comments = WorkloadAllocation.query.join(User, WorkloadAllocation.username == User.username).filter(
            User.dept_id == hod_dept_id
        ).all()

        # Render the comment history page
        return render_template('comment_history.html', title='Comment History', comments=comments)

    # If the current user is not an HoD, redirect them to some other page or display an error message (depends on your application's requirements)
    else:
        flash('You do not have permission to view the Comment History page.')
        return redirect(url_for('dashboard'))


@app.route('/mark_comments_as_read', methods=['POST'])
@login_required
def mark_comments_as_read():
    # Check if the user is an HoD
    if g.is_hod:
        # Fetch the department ID of the current HoD user
        hod_dept_id = User.query.filter_by(
            username=current_user.username).first().dept_id

        # Fetch the comments made by staff members with the same dept_id as that of the HoD user
        comments = WorkloadAllocation.query.join(User, WorkloadAllocation.username == User.username).filter(
            User.dept_id == hod_dept_id
        ).all()

        # Return success status as JSON response
        return jsonify(status="success"), 200

    # If the user is not an HoD, return an error message
    else:
        return jsonify(status="error", message="Unauthorised access"), 403

# Update comment_status from Unread to Read upon clicking the "Read & Close" button in the modal of the Comment History page
# The URL of this route has a dynamic part 'alloc_id' which is an integer.


@app.route('/update_comment_status/<int:alloc_id>', methods=['POST'])
@login_required
def update_comment_status(alloc_id):
    # Fetch comment from the database using the provided 'alloc_id'
    comment = WorkloadAllocation.query.get(alloc_id)

    # Check if the comment exists
    if comment:
        if comment.comment_status == "Unread":
            comment.comment_status = "Read"
            db.session.commit()
            # Return a success status with a 200 OK HTTP status code
            return jsonify(status="success"), 200
        else:
            # If the comment is already "Read", return a status indicating that it's already read with a 200 OK HTTP status code.
            return jsonify(status="alreadyRead"), 200
    # If the comment doesn't exist, return an error status with a 404 Not Found HTTP status code.
    return jsonify(status="error", message="Comment not found"), 404


# The below function runs before processing a request


@app.before_request
def before_request():
    # Initialise the global notification flag to False
    g.show_notification = False
    # Check if the current user is authenticated
    if current_user.is_authenticated:
        # Fetch the user data from the User table based on the current user's username
        user = User.query.filter_by(username=current_user.username).first()

        # If the user is found
        if user:
            # Check if the user's role is "HoD"
            is_hod = user.role_id == Role.query.filter_by(
                role_name='HoD').first().role_id

            # Set a global variable to identify if the current user is an HoD
            g.is_hod = is_hod

            # If the user is an HoD
            if is_hod:
                # Query for unread comments from users in the same department as the HoD
                unread_comments = WorkloadAllocation.query.join(User, WorkloadAllocation.username == User.username).filter(
                    WorkloadAllocation.comment_status == "Unread",
                    User.dept_id == user.dept_id
                ).count()

                # If there are any unread comments, set the global notification flag to True
                g.show_notification = unread_comments > 0
            else:
                # If user is not an HoD, set the global notification flag to False
                g.show_notification = False
        else:
            # If user data is not found in the User table, set the global notification flag to False
            g.show_notification = False

# Check for unread comments


@app.route('/check_unread_comments')
@login_required  # Ensure that the user is logged in to access this route
def check_unread_comments():
    # Fetch the user data from the User table based on the current user's username
    user = User.query.filter_by(username=current_user.username).first()

    # If the user is found
    if user:
        # Query for unread comments from users in the same department as the current user
        unread_comments = WorkloadAllocation.query.join(User, WorkloadAllocation.username == User.username).filter(
            WorkloadAllocation.comment_status == "Unread",
            User.dept_id == user.dept_id
        ).count()
    else:
        # If user data is not found, set unread comments to 0
        unread_comments = 0

    # Return the number of unread comments as a JSON response
    return jsonify(unread_comments=unread_comments)


if __name__ == '__main__':
    app.run(debug=True)
